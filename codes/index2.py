import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def load_json_data(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

def create_sheet(wb, title, headers):
    sheet = wb.create_sheet(title=title)
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    return sheet

def populate_author_info(sheet, data):
    author = data['Data']['Records'][0]
    row = [
        author['author_name'],
        author['authorAliasName'],
        author['publishingName'],
        author['rid'],
        author['orcid'],
        author['authorId'],
        author['institution'],
        author['primaryAffiliation'],
        author['primaryAffiliationLocation'],
        ', '.join(author['primaryAffiliationDepartment']),
        author['photoUrlLarge']
    ]
    sheet.append(row)

def populate_publication_summary(sheet, data):
    summary = data['Data']['Records'][0]['publication_summary']
    row = [summary[key] for key in summary]
    sheet.append(row)

def populate_journals(sheet, data):
    journals = data['Data']['Records'][0]['fullJournalsList']
    for journal in journals:
        sheet.append([journal['value'], journal['count']])

def populate_affiliations(sheet, data):
    affiliations = data['Data']['Records'][0]['affiliations']
    for affiliation in affiliations:
        sheet.append([affiliation['value'], affiliation['count'], affiliation['minYear'], affiliation['maxYear']])

def populate_categories(sheet, data):
    categories = data['Data']['Records'][0]['categories']
    for category in categories:
        sheet.append([category])

def populate_alternative_names(sheet, data):
    alt_names = data['Data']['Records'][0]['alternativeName']
    for name in alt_names:
        sheet.append([name['value'], name['count']])

def populate_citation_stats(sheet, data):
    wos_stats = data['Data']['Records'][0]['wos_stats']
    diidw_stats = data['Data']['Records'][0]['diidw_stats']
    
    for key, value in wos_stats.items():
        sheet.append([f"WoS {key.replace('_', ' ').title()}", value])
    
    for key, value in diidw_stats.items():
        sheet.append([f"DIIDW {key.replace('_', ' ').title()}", value])

def populate_yearly_citations(sheet, data):
    citations = data['Data']['Records'][0]['publonsBasicMetricsData']['publonsIndividualStatsData']['citationsPerYear']
    for year, count in citations.items():
        sheet.append([year, int(count)])

def populate_publons_stats(sheet, data):
    stats = data['Data']['Records'][0]['publonsBasicMetricsData']['publonsStatData']
    for key in stats['total'].keys():
        sheet.append([
            key,
            stats['total'][key],
            stats['median'][key],
            stats['percentile'][key]
        ])

def main():
    data = load_json_data('banalaxmi.json')
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create and populate sheets
    populate_author_info(create_sheet(wb, "Author Information", ["Author Name", "Alias Name", "Publishing Name", "RID", "ORCID", "Author ID", "Institution", "Primary Affiliation", "Primary Affiliation Location", "Primary Affiliation Department", "Photo URL"]), data)
    populate_publication_summary(create_sheet(wb, "Publication Summary", ["Total", "Indexed", "Non-Indexed", "WOSCC", "PPRN", "PQDT"]), data)
    populate_journals(create_sheet(wb, "Journals", ["Journal Name", "Publication Count"]), data)
    populate_affiliations(create_sheet(wb, "Affiliations", ["Affiliation Name", "Count", "Min Year", "Max Year"]), data)
    populate_categories(create_sheet(wb, "Categories", ["Category Name"]), data)
    populate_alternative_names(create_sheet(wb, "Alternative Names", ["Alternative Name", "Count"]), data)
    populate_citation_stats(create_sheet(wb, "Citation Statistics", ["Metric Name", "Value"]), data)
    populate_yearly_citations(create_sheet(wb, "Yearly Citations", ["Year", "Citation Count"]), data)
    populate_publons_stats(create_sheet(wb, "Publons Stats", ["Stat Name", "Total Value", "Median Value", "Percentile"]), data)
    
    # Save the workbook
    wb.save("banalaxmi.xlsx")
    print("Excel file 'author_data.xlsx' has been created successfully.")

if __name__ == "__main__":
    main()
